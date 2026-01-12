"""Service for report generation and management."""

import io
import logging
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy.ext.asyncio import AsyncSession

from core.service import BaseService
from dao import DashboardDAO, ReportDAO
from dto.dashboard import RequestCountDTO
from models.ambulance_request import RequestStatus
from models.report import ReportFormat
from schemas.report import (
    GenerateReportResponseSchema,
    LatestReportsResponseSchema,
    ReportItemSchema,
    ReportStatisticsSchema,
)
from services.aws.actions import S3Actions
from services.dashboard_metrics.metrics_calculator import (
    DashboardMetricsCalculator,
)

logger = logging.getLogger(__name__)


class ReportService(BaseService):
    """Service for report generation and management."""

    def __init__(
        self,
        db_session: AsyncSession,
        *,
        report_dao: ReportDAO | None = None,
        dashboard_dao: DashboardDAO | None = None,
        s3_actions: S3Actions | None = None,
    ):
        """Initialize ReportService.

        Args:
            db_session: Database session.
            report_dao: Optional ReportDAO instance.
            dashboard_dao: Optional DashboardDAO instance.
            s3_actions: Optional S3Actions instance.

        """
        super().__init__(db_session)
        self._report_dao = report_dao or ReportDAO(db_session)
        self._dashboard_dao = dashboard_dao or DashboardDAO(db_session)
        self._s3_actions = s3_actions or S3Actions()

    def _validate_period_dates(
        self,
        start_date: date,
        end_date: date,
    ) -> None:
        """Validate period start and end dates.

        Args:
            start_date: Start date.
            end_date: End date.

        Raises:
            ValueError: If period parameters are invalid.

        """
        if start_date > end_date:
            raise ValueError('start_date must be <= end_date')  # noqa: TRY003

    async def _get_current_statistics(self) -> RequestCountDTO:
        """Get current statistics for all requests.

        Returns:
            RequestCountDTO with current statistics.

        """
        counts = await self._dashboard_dao.get_request_counts_by_status()
        return RequestCountDTO(
            approved_all=counts.get(RequestStatus.APPROVED, 0),
            pending_all=counts.get(RequestStatus.SUBMITTED, 0),
            denied_all=counts.get(RequestStatus.DENIED, 0),
        )

    def _generate_report_name(
        self, file_format: ReportFormat, period_start: date, period_end: date
    ) -> str:
        """Generate report name.

        Args:
            file_format: Report format.
            period_start: Period start date.
            period_end: Period end date.

        Returns:
            Report name string.

        """
        date_str = f'{period_start.strftime("%Y-%m-%d")}_to_{period_end.strftime("%Y-%m-%d")}'  # noqa: E501
        return f'Report_{date_str}.{file_format.value}'

    def _generate_pdf_report(
        self,
        period_start: date,
        period_end: date,
        statistics: RequestCountDTO,
    ) -> bytes:
        """Generate PDF report.

        Args:
            period_start: Period start date.
            period_end: Period end date.
            statistics: Request statistics.

        Returns:
            PDF file bytes.

        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story: list[Any] = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph(
            f'Report: {period_start} to {period_end}',
            styles['Title'],
        )
        story.append(title)
        story.append(Spacer(1, 0.3 * inch))

        # Statistics table
        data = [
            ['Metric', 'Value'],
            ['Total Requests', str(statistics.total_requests)],
            ['Approved Requests', str(statistics.approved_all)],
            ['Denied Requests', str(statistics.denied_all)],
            ['Pending Requests', str(statistics.pending_all)],
            ['Approval Rate', f'{statistics.approval_rate:.2f}%'],
            ['Denial Rate', f'{statistics.denial_rate:.2f}%'],
        ]

        table = Table(data, colWidths=[3 * inch, 2 * inch])
        table.setStyle(
            TableStyle(
                [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(table)

        doc.build(story)
        buffer.seek(0)
        return buffer.read()

    def _generate_excel_report(
        self,
        period_start: date,
        period_end: date,
        statistics: RequestCountDTO,
    ) -> bytes:
        """Generate Excel report.

        Args:
            period_start: Period start date.
            period_end: Period end date.
            statistics: Request statistics.

        Returns:
            Excel file bytes.

        """
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError('Failed to get active worksheet')  # noqa: TRY003

        ws.title = 'Report'

        # Title
        ws['A1'] = f'Report: {period_start} to {period_end}'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        # Statistics
        headers = ['Metric', 'Value']
        ws.append(headers)

        # Format header row
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left')

        data: list[list[str | int | float]] = [
            ['Total Requests', statistics.total_requests],
            ['Approved Requests', statistics.approved_all],
            ['Denied Requests', statistics.denied_all],
            ['Pending Requests', statistics.pending_all],
            ['Approval Rate', f'{statistics.approval_rate:.2f}%'],
            ['Denial Rate', f'{statistics.denial_rate:.2f}%'],
        ]

        for row in data:
            ws.append(row)

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    async def generate_report(
        self,
        *,
        file_format: ReportFormat,
        start_date: date,
        end_date: date,
        created_by_id: int,
    ) -> GenerateReportResponseSchema:
        """Generate a report.

        Args:
            file_format: Report format (PDF or Excel).
            start_date: Start date for report period.
            end_date: End date for report period.
            created_by_id: ID of the user creating the report.

        Returns:
            GenerateReportResponseSchema with report ID and download URL.

        Raises:
            ValueError: If start_date > end_date.

        """
        self._validate_period_dates(start_date, end_date)

        # Get current statistics
        statistics = await self._get_current_statistics()

        # Generate report file
        if file_format == ReportFormat.PDF:
            file_bytes = self._generate_pdf_report(
                start_date, end_date, statistics
            )
            file_name = self._generate_report_name(
                file_format, start_date, end_date
            )
            content_type = 'application/pdf'
        else:  # Excel
            file_bytes = self._generate_excel_report(
                start_date, end_date, statistics
            )
            file_name = self._generate_report_name(
                file_format, start_date, end_date
            )
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # noqa: E501

        # Upload to S3
        file_obj = io.BytesIO(file_bytes)
        s3_key, _ = self._s3_actions.upload_file(
            file_obj=file_obj,
            file_name=file_name,
            file_size=len(file_bytes),
            declared_content_type=content_type,
            prefix='reports',
        )

        # Create report record
        report_name = self._generate_report_name(
            file_format, start_date, end_date
        )
        report = await self._report_dao.create(
            name=report_name,
            file_format=file_format.value,
            s3_key=s3_key,
            created_by_id=created_by_id,
            period_start=start_date,
            period_end=end_date,
            total_requests=statistics.total_requests,
            approved_requests=statistics.approved_all,
            denied_requests=statistics.denied_all,
            pending_requests=statistics.pending_all,
        )
        await self._session.commit()

        # Generate presigned URL for download
        download_url = self._s3_actions.get_presigned_url(key=s3_key)

        return GenerateReportResponseSchema(
            report_id=report.id,
            download_url=download_url,
        )

    async def get_latest_reports_with_statistics(
        self,
    ) -> LatestReportsResponseSchema:
        """Get latest reports with current statistics comparison.

        Returns:
            LatestReportsResponseSchema with reports and statistics.

        """
        # Get latest reports
        reports = await self._report_dao.get_latest_reports(limit=3)

        # Get current statistics
        current_stats = await self._get_current_statistics()

        # Get latest report for comparison
        latest_report = await self._report_dao.get_latest_report()

        # Calculate changes
        if latest_report:
            change_total = (
                DashboardMetricsCalculator.calculate_percentage_change(
                    current_value=current_stats.total_requests,
                    previous_value=latest_report.total_requests,
                )
            )
            change_approved = (
                DashboardMetricsCalculator.calculate_percentage_change(
                    current_value=current_stats.approved_all,
                    previous_value=latest_report.approved_requests,
                )
            )
            change_denied = (
                DashboardMetricsCalculator.calculate_percentage_change(
                    current_value=current_stats.denied_all,
                    previous_value=latest_report.denied_requests,
                )
            )
            change_pending = (
                DashboardMetricsCalculator.calculate_percentage_change(
                    current_value=current_stats.pending_all,
                    previous_value=latest_report.pending_requests,
                )
            )
        else:
            change_total = 0.0
            change_approved = 0.0
            change_denied = 0.0
            change_pending = 0.0

        statistics = ReportStatisticsSchema(
            total_requests=current_stats.total_requests,
            total_requests_change=change_total,
            approved_requests=current_stats.approved_all,
            approved_requests_change=change_approved,
            denied_requests=current_stats.denied_all,
            denied_requests_change=change_denied,
            pending_requests=current_stats.pending_all,
            pending_requests_change=change_pending,
        )

        report_items = [
            ReportItemSchema(
                id=report.id,
                name=report.name,
                format=ReportFormat(report.format),
                created_at=report.created_at,
                created_by_full_name=f'{report.created_by.name} '
                f'{report.created_by.surname}',
                s3_key=report.s3_key,
            )
            for report in reports
        ]

        return LatestReportsResponseSchema(
            reports=report_items,
            current_statistics=statistics,
        )
